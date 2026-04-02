#!/usr/bin/env python3
"""Fetch Nord Pool day-ahead prices and save them into an Excel workbook."""

from __future__ import annotations

from datetime import datetime, date
from pathlib import Path
from typing import Dict, List


# =====================================================
# Set your target Excel file path here before running.
# =====================================================
EXCEL_PATH = r"/workspace/CBC-Spreads/nordpool_prices.xlsx"

API_URL = "https://dataportal-api.nordpoolgroup.com/api/DayAheadPriceIndices"
INDEX_NAMES = [
    "EE",
    "LT",
    "LV",
    "AT",
    "BE",
    "FR",
    "GER",
    "NL",
    "PL",
    "DK1",
    "DK2",
    "FI",
    "NO1",
    "NO2",
    "NO3",
    "NO4",
    "NO5",
    "SE1",
    "SE2",
    "SE3",
    "SE4",
    "BG",
    "TEL",
]

# Pairs used for cross-border spread table (left_area -> right_area).
# Spread = right_area_price - left_area_price
NEIGHBOUR_PAIRS = [
    ("EE", "LT"),
    ("LT", "LV"),
    ("AT", "BE"),
    ("BE", "FR"),
    ("FR", "GER"),
    ("GER", "NL"),
    ("NL", "PL"),
    ("DK1", "DK2"),
    ("NO1", "NO2"),
    ("NO2", "NO3"),
    ("NO3", "NO4"),
    ("NO4", "NO5"),
    ("SE1", "SE2"),
    ("SE2", "SE3"),
    ("SE3", "SE4"),
    ("BG", "TEL"),
]


def show_menu() -> str:
    inner_width = 42

    def box_line(content: str = "") -> str:
        return f"║{content:<{inner_width}}║"

    menu_lines = [
        "",
        f"╔{'═' * inner_width}╗",
        box_line("   Nord Pool RO (TEL)"),
        f"╠{'═' * inner_width}╣",
        box_line(),
        box_line("  [1]  Mai adatok (Today)"),
        box_line("  [2]  Konkrét dátum (pl. 04-01)"),
        box_line("  [0]  Kilépés"),
        box_line(),
        f"╚{'═' * inner_width}╝",
        "",
        "Válassz [1/2/0]: ",
    ]

    return input("\n".join(menu_lines)).strip()


def parse_target_date() -> date | None:
    while True:
        choice = show_menu()

        if choice == "0":
            return None

        if choice == "1":
            return datetime.now().date()

        if choice == "2":
            raw = input("Add meg a dátumot (MM-DD vagy YYYY-MM-DD): ").strip()
            for fmt in ("%Y-%m-%d", "%m-%d"):
                try:
                    parsed = datetime.strptime(raw, fmt)
                    if fmt == "%m-%d":
                        parsed = parsed.replace(year=datetime.now().year)
                    return parsed.date()
                except ValueError:
                    continue

            print("Hibás dátum formátum. Próbáld újra.")
            continue

        print("Érvénytelen választás. Kérlek válassz 1, 2 vagy 0 opciót.")


def fetch_prices(target_date: date) -> Dict:
    import requests
    params = {
        "date": target_date.isoformat(),
        "market": "DayAhead",
        "indexNames": ",".join(INDEX_NAMES),
        "currency": "EUR",
        "resolutionInMinutes": 60,
    }

    headers = {
        "accept": "application/json, text/plain, */*",
        "origin": "https://data.nordpoolgroup.com",
        "referer": "https://data.nordpoolgroup.com/",
        "user-agent": "Mozilla/5.0",
    }

    response = requests.get(API_URL, params=params, headers=headers, timeout=30)
    response.raise_for_status()
    return response.json()


def build_rows(payload: Dict) -> List[List[float | str | None]]:
    rows: List[List[float | str | None]] = []
    entries = payload.get("multiIndexEntries", [])

    for idx, item in enumerate(entries, start=1):
        entry_per_area = item.get("entryPerArea", {})
        row: List[float | str | None] = [idx]
        for area in INDEX_NAMES:
            row.append(entry_per_area.get(area))
        rows.append(row)

    return rows


def save_to_excel(target_date: date, rows: List[List[float | str | None]]) -> None:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    excel_path = Path(EXCEL_PATH)
    excel_path.parent.mkdir(parents=True, exist_ok=True)

    if excel_path.exists():
        wb = load_workbook(excel_path)
    else:
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

    sheet_name = str(target_date.day)
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])

    ws = wb.create_sheet(title=sheet_name)

    header_fill = PatternFill("solid", fgColor="BDD7EE")
    data_fill = PatternFill("solid", fgColor="FCE4D6")
    avg_fill = PatternFill("solid", fgColor="FFFF00")
    spread_fill = PatternFill("solid", fgColor="E2F0D9")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    header_font = Font(bold=True)

    headers = ["Hour"] + INDEX_NAMES
    ws.append(headers)
    for row in rows:
        ws.append(row)

    avg_row = len(rows) + 2
    ws.cell(row=avg_row, column=1, value="AVG")
    for col in range(2, len(headers) + 1):
        col_letter = get_column_letter(col)
        ws.cell(
            row=avg_row,
            column=col,
            value=f"=AVERAGE({col_letter}2:{col_letter}{avg_row - 1})",
        )

    for row_idx in range(1, avg_row + 1):
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = center
            cell.border = border

            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
            elif row_idx == avg_row:
                cell.fill = avg_fill
                cell.font = header_font
            else:
                cell.fill = data_fill

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 10

    spread_start_row = avg_row + 3
    spread_headers = ["Hour"] + [f"{left}-{right}" for left, right in NEIGHBOUR_PAIRS]
    ws.append([])
    ws.append([])
    ws.append(spread_headers)

    for hour_idx in range(1, len(rows) + 1):
        excel_row = spread_start_row + hour_idx
        ws.cell(row=excel_row, column=1, value=hour_idx)
        for col_idx, (left, right) in enumerate(NEIGHBOUR_PAIRS, start=2):
            left_col = INDEX_NAMES.index(left) + 2
            right_col = INDEX_NAMES.index(right) + 2
            ws.cell(
                row=excel_row,
                column=col_idx,
                value=(
                    f"={get_column_letter(right_col)}{hour_idx + 1}"
                    f"-{get_column_letter(left_col)}{hour_idx + 1}"
                ),
            )

    spread_avg_row = spread_start_row + len(rows) + 1
    ws.cell(row=spread_avg_row, column=1, value="AVG")
    for col_idx in range(2, len(spread_headers) + 1):
        col_letter = get_column_letter(col_idx)
        ws.cell(
            row=spread_avg_row,
            column=col_idx,
            value=f"=AVERAGE({col_letter}{spread_start_row + 1}:{col_letter}{spread_avg_row - 1})",
        )

    for row_idx in range(spread_start_row, spread_avg_row + 1):
        for col_idx in range(1, len(spread_headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = center
            cell.border = border

            if row_idx == spread_start_row:
                cell.fill = header_fill
                cell.font = header_font
            elif row_idx == spread_avg_row:
                cell.fill = avg_fill
                cell.font = header_font
            else:
                cell.fill = spread_fill

    for col_idx in range(1, len(spread_headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(
            ws.column_dimensions[get_column_letter(col_idx)].width or 10,
            11,
        )

    wb.save(excel_path)


def main() -> None:
    target_date = parse_target_date()
    if target_date is None:
        print("Kilépés...")
        return

    try:
        payload = fetch_prices(target_date)
    except Exception as exc:
        print(f"Hálózati/API hiba történt: {exc}")
        return

    rows = build_rows(payload)

    if not rows:
        print(f"Nincs adat a következő dátumra: {target_date.isoformat()}")
        return

    save_to_excel(target_date, rows)
    print(
        f"Sikeres mentés: {len(rows)} óra adat -> '{EXCEL_PATH}', munkalap: '{target_date.day}'"
    )


if __name__ == "__main__":
    main()
