#!/usr/bin/env python3
"""Fetch Nord Pool day-ahead prices and save them into an Excel workbook."""

from __future__ import annotations

import json
import urllib.parse
import urllib.request
from datetime import datetime, date, timedelta
from pathlib import Path
from typing import Dict, List


# =====================================================
# Set your target Excel file path here before running.
# =====================================================
EXCEL_PATH = r"C:\Users\Lucas\Desktop\NordPoolDamSpreads.xlsx"

API_URL = "https://dataportal-api.nordpoolgroup.com/api/DayAheadPriceIndices"
# HUPX Labs API base URL (used for HU DAM prices)
BASE_URL = "https://labs.hupx.hu/data/v1"

# Stripped down strictly to the requested countries (DE changed to GER)
INDEX_NAMES = [
    "EE",
    "LV",
    "AT",
    "BE",
    "FR",
    "GER",
    "NL",
    "PL",
    "DK1",
    "DK2",
    "FI",
    "HU",
    "BG",
    "TEL",
]

# Pairs imported from country_pairs.json with corrected designations (GER, TEL, DK1)
# Spread = right_area_price - left_area_price
NEIGHBOUR_PAIRS = [
    ("AT", "GER"),
    ("AT", "HU"),
    ("BE", "FR"),
    ("BE", "NL"),
    ("BG", "TEL"),
    ("DK1", "DK2"),
    ("DK1", "GER"),
    ("DK2", "GER"),
    ("GER", "FR"),
    ("GER", "NL"),
    ("DK1", "NL"),
    ("EE", "LV"),
    ("FI", "EE"),
    ("HU", "TEL"),
    ("TEL", "HU"),
    ("PL", "GER"),
    ("PL", "NL")
]

def show_menu() -> str:
    inner_width = 42

    def box_line(content: str = "") -> str:
        return f"║{content:<{inner_width}}║"

    menu_lines = [
        "",
        f"╔{'═' * inner_width}╗",
        box_line("   CBC DAM Spread "),
        f"╠{'═' * inner_width}╣",
        box_line(),
        box_line("  [1]  Mai adatok (Today)"),
        box_line("  [2]  Konkrét dátum (pl. 04-01)"),
        box_line("  [0]  Kilépés"),
        box_line(),
        f"╚{'═' * inner_width}╝",
        "",
        "Válassz [1/2/0]: ",
    ]

    return input("\n".join(menu_lines)).strip()


def parse_target_date() -> date | None:
    while True:
        choice = show_menu()

        if choice == "0":
            return None

        if choice == "1":
            return datetime.now().date()

        if choice == "2":
            raw = input("Add meg a dátumot (MM-DD vagy YYYY-MM-DD): ").strip()
            for fmt in ("%Y-%m-%d", "%m-%d"):
                try:
                    parsed = datetime.strptime(raw, fmt)
                    if fmt == "%m-%d":
                        parsed = parsed.replace(year=datetime.now().year)
                    return parsed.date()
                except ValueError:
                    continue

            print("Hibás dátum formátum. Próbáld újra.")
            continue

        print("Érvénytelen választás. Kérlek válassz 1, 2 vagy 0 opciót.")


def fetch_prices(target_date: date) -> Dict:
    import requests
    params = {
        "date": target_date.isoformat(),
        "market": "DayAhead",
        "indexNames": ",".join(INDEX_NAMES),
        "currency": "EUR",
        "resolutionInMinutes": 60,
    }

    headers = {
        "accept": "application/json, text/plain, */*",
        "origin": "https://data.nordpoolgroup.com",
        "referer": "https://data.nordpoolgroup.com/",
        "user-agent": "Mozilla/5.0",
    }

    response = requests.get(API_URL, params=params, headers=headers, timeout=30)
    response.raise_for_status()
    return response.json()


def fetch_json(endpoint: str, filters: List[str], limit: int = 200) -> List[Dict]:
    filter_str = ",".join(filters)
    url = f"{BASE_URL}/{endpoint}?filter={urllib.parse.quote(filter_str)}&limit={limit}"
    all_data: List[Dict] = []

    while url:
        req = urllib.request.Request(url)
        req.add_header("User-Agent", "HUPX-Fetcher/1.0")
        with urllib.request.urlopen(req, timeout=30) as resp:
            body = json.loads(resp.read().decode())

        all_data.extend(body.get("data", []))
        url = body.get("nextPage")

    return all_data


def fetch_dam(date_str: str) -> Dict[int, Dict[str, float | None]]:
    """Returns {hour: {price, volume}} for DAM HU."""
    next_date = (datetime.strptime(date_str, "%Y-%m-%d") + timedelta(days=1)).strftime(
        "%Y-%m-%d"
    )
    rows = fetch_json(
        "dam_aggregated_trading_data",
        [
            f"DeliveryDay__gte__{date_str}",
            f"DeliveryDay__lt__{next_date}",
            "Region__eq__HU",
        ],
    )

    result: Dict[int, Dict[str, float | None]] = {}
    for r in rows:
        hour = int(r["ProductH"])
        result[hour] = {"price": r.get("Price"), "volume": r.get("Volume")}

    return result


def build_rows(payload: Dict, hu_dam: Dict[int, Dict[str, float | None]]) -> List[List[float | str | None]]:
    rows: List[List[float | str | None]] = []
    entries = payload.get("multiIndexEntries", [])

    for idx, item in enumerate(entries, start=1):
        entry_per_area = item.get("entryPerArea", {})
        hu_price = hu_dam.get(idx, {}).get("price")
        entry_per_area["HU"] = hu_price
        row: List[float | str | None] = [idx]
        for area in INDEX_NAMES:
            row.append(entry_per_area.get(area))
        rows.append(row)

    return rows


def save_to_excel(target_date: date, rows: List[List[float | str | None]]) -> None:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    excel_path = Path(EXCEL_PATH)
    excel_path.parent.mkdir(parents=True, exist_ok=True)

    if excel_path.exists():
        wb = load_workbook(excel_path)
    else:
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

    sheet_name = target_date.strftime("%Y.%m.%d.")
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])

    ws = wb.create_sheet(title=sheet_name)

    header_fill = PatternFill("solid", fgColor="BDD7EE")
    data_fill = PatternFill("solid", fgColor="FCE4D6")
    avg_fill = PatternFill("solid", fgColor="FFFF00")
    spread_fill = PatternFill("solid", fgColor="E2F0D9")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    header_font = Font(bold=True)

    headers = ["Hour"] + INDEX_NAMES
    ws.append(headers)
    for row in rows:
        ws.append(row)

    avg_row = len(rows) + 2
    ws.cell(row=avg_row, column=1, value="AVG")
    for col in range(2, len(headers) + 1):
        col_letter = get_column_letter(col)
        ws.cell(
            row=avg_row,
            column=col,
            value=f"=AVERAGE({col_letter}2:{col_letter}{avg_row - 1})",
        )

    for row_idx in range(1, avg_row + 1):
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = center
            cell.border = border

            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
            elif row_idx == avg_row:
                cell.fill = avg_fill
                cell.font = header_font
            else:
                cell.fill = data_fill

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 10

    spread_start_row = avg_row + 3
    spread_headers = ["Hour"] + [f"{left}-{right}" for left, right in NEIGHBOUR_PAIRS]
    ws.append([])
    ws.append([])
    ws.append(spread_headers)

    for hour_idx in range(1, len(rows) + 1):
        excel_row = spread_start_row + hour_idx
        ws.cell(row=excel_row, column=1, value=hour_idx)
        for col_idx, (left, right) in enumerate(NEIGHBOUR_PAIRS, start=2):
            try:
                left_col = INDEX_NAMES.index(left) + 2
                right_col = INDEX_NAMES.index(right) + 2
                ws.cell(
                    row=excel_row,
                    column=col_idx,
                    value=(
                        f"={get_column_letter(right_col)}{hour_idx + 1}"
                        f"-{get_column_letter(left_col)}{hour_idx + 1}"
                    ),
                )
            except ValueError:
                pass

    spread_avg_row = spread_start_row + len(rows) + 1
    ws.cell(row=spread_avg_row, column=1, value="AVG")
    for col_idx in range(2, len(spread_headers) + 1):
        col_letter = get_column_letter(col_idx)
        ws.cell(
            row=spread_avg_row,
            column=col_idx,
            value=f"=AVERAGE({col_letter}{spread_start_row + 1}:{col_letter}{spread_avg_row - 1})",
        )

    for row_idx in range(spread_start_row, spread_avg_row + 1):
        for col_idx in range(1, len(spread_headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = center
            cell.border = border

            if row_idx == spread_start_row:
                cell.fill = header_fill
                cell.font = header_font
            elif row_idx == spread_avg_row:
                cell.fill = avg_fill
                cell.font = header_font
            else:
                cell.fill = spread_fill

    for col_idx in range(1, len(spread_headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(
            ws.column_dimensions[get_column_letter(col_idx)].width or 10,
            11,
        )

    wb.save(excel_path)


def main() -> None:
    target_date = parse_target_date()
    if target_date is None:
        print("Kilépés...")
        return

    try:
        payload = fetch_prices(target_date)
    except Exception as exc:
        print(f"Hálózati/API hiba történt: {exc}")
        return

    hu_dam: Dict[int, Dict[str, float | None]] = {}
    try:
        hu_dam = fetch_dam(target_date.isoformat())
    except Exception as exc:
        print(f"HU DAM adatlekérés sikertelen, HU oszlop üres marad: {exc}")

    rows = build_rows(payload, hu_dam)

    if not rows:
        print(f"Nincs adat a következő dátumra: {target_date.isoformat()}")
        return

    save_to_excel(target_date, rows)
    print(
        f"Sikeres mentés: {len(rows)} óra adat -> '{EXCEL_PATH}', munkalap: '{target_date.strftime('%Y.%m.%d.')}'"
    )


if __name__ == "__main__":
    main()
